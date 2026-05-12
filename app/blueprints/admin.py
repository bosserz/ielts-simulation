import json
import uuid
from functools import wraps
from flask import Blueprint, render_template, abort, redirect, url_for, flash, request, jsonify
from flask_login import login_required, current_user
from ..models.user import User, UserRole
from ..models.exam import Exam, ExamStatus, ExamType, Section, SectionType, Question, QuestionType
from ..models.session import ExamSession, SessionStatus
from ..extensions import db

admin_bp = Blueprint("admin", __name__, url_prefix="/admin")


def teacher_required(f):
    @wraps(f)
    def decorated(*args, **kwargs):
        if not current_user.is_authenticated or not current_user.is_teacher:
            abort(403)
        return f(*args, **kwargs)
    return decorated


@admin_bp.route("/dashboard")
@login_required
@teacher_required
def dashboard():
    total_students = User.query.filter_by(role=UserRole.STUDENT).count()
    recent_sessions = (
        ExamSession.query
        .filter(ExamSession.status.in_([SessionStatus.SUBMITTED, SessionStatus.SCORED]))
        .order_by(ExamSession.submitted_at.desc())
        .limit(10)
        .all()
    )
    return render_template(
        "admin/dashboard.html",
        total_students=total_students,
        recent_sessions=recent_sessions,
    )


@admin_bp.route("/students")
@login_required
@teacher_required
def students():
    query = User.query.filter_by(role=UserRole.STUDENT)
    users = query.order_by(User.name).all()
    return render_template("admin/students.html", users=users)


@admin_bp.route("/students/<user_id>")
@login_required
@teacher_required
def student_profile(user_id: str):
    from .student import _compute_readiness
    student = User.query.filter_by(id=user_id, role=UserRole.STUDENT).first_or_404()
    sessions = (
        ExamSession.query
        .filter_by(user_id=user_id)
        .order_by(ExamSession.submitted_at.desc())
        .all()
    )
    published_exams = Exam.query.filter_by(status=ExamStatus.PUBLISHED).order_by(Exam.title).all()
    readiness = _compute_readiness(student, sessions)
    return render_template(
        "admin/student_profile.html",
        student=student,
        sessions=sessions,
        published_exams=published_exams,
        readiness=readiness,
    )


@admin_bp.route("/students/<user_id>/assign", methods=["POST"])
@login_required
@teacher_required
def assign_session(user_id: str):
    student = User.query.filter_by(id=user_id, role=UserRole.STUDENT).first_or_404()
    exam_id = request.form.get("exam_id")
    if not exam_id:
        flash("Please select an exam.", "error")
        return redirect(url_for("admin.student_profile", user_id=user_id))

    session = ExamSession(exam_id=exam_id, user_id=student.id)
    db.session.add(session)
    db.session.commit()
    flash(f"Exam assigned to {student.name}.", "success")
    return redirect(url_for("admin.student_profile", user_id=user_id))


@admin_bp.route("/exams")
@login_required
@teacher_required
def exams():
    all_exams = Exam.query.order_by(Exam.created_at.desc()).all()
    return render_template("admin/exams.html", exams=all_exams)


@admin_bp.route("/exams/<exam_id>/toggle-status", methods=["POST"])
@login_required
@teacher_required
def toggle_exam_status(exam_id: str):
    exam = Exam.query.filter_by(id=exam_id).first_or_404()
    action = request.form.get("action")
    if action == "publish":
        exam.status = ExamStatus.PUBLISHED
    elif action == "archive":
        exam.status = ExamStatus.ARCHIVED
    db.session.commit()
    flash(f"'{exam.title}' status updated.", "success")
    return redirect(url_for("admin.exams"))


# ---------------------------------------------------------------------------
# Exam Builder
# ---------------------------------------------------------------------------

@admin_bp.route("/exams/new", methods=["GET", "POST"])
@login_required
@teacher_required
def exam_new():
    if request.method == "POST":
        title = request.form.get("title", "").strip()
        exam_type = request.form.get("type", ExamType.ACADEMIC)
        if not title:
            flash("Title is required.", "error")
            return redirect(url_for("admin.exam_new"))
        exam = Exam(title=title, type=exam_type, created_by=current_user.id)
        db.session.add(exam)
        db.session.commit()
        flash(f"Exam '{title}' created.", "success")
        return redirect(url_for("admin.exam_edit", exam_id=exam.id))
    return render_template("admin/exam_new.html")


@admin_bp.route("/exams/<exam_id>/edit")
@login_required
@teacher_required
def exam_edit(exam_id: str):
    exam = Exam.query.filter_by(id=exam_id).first_or_404()
    section_types = [SectionType.LISTENING, SectionType.READING, SectionType.WRITING, SectionType.SPEAKING]
    question_types = [
        QuestionType.MCQ, QuestionType.TFNG, QuestionType.YNGNG,
        QuestionType.SHORT_ANSWER, QuestionType.NOTE_COMPLETION,
        QuestionType.SENTENCE_COMPLETION, QuestionType.MATCHING_HEADINGS,
        QuestionType.MATCHING_FEATURES, QuestionType.TABLE_COMPLETION,
        QuestionType.SUMMARY_COMPLETION,
    ]
    return render_template(
        "admin/exam_builder.html",
        exam=exam,
        section_types=section_types,
        question_types=question_types,
    )


@admin_bp.route("/exams/<exam_id>/sections", methods=["POST"])
@login_required
@teacher_required
def section_add(exam_id: str):
    exam = Exam.query.filter_by(id=exam_id).first_or_404()
    sec_type = request.form.get("type", SectionType.LISTENING)
    time_limit = int(request.form.get("time_limit_min") or 30) * 60
    config_raw = request.form.get("config", "{}")
    try:
        config = json.loads(config_raw) if config_raw.strip() else {}
    except ValueError:
        flash("Config must be valid JSON.", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))
    order = len(exam.sections) + 1
    section = Section(exam_id=exam_id, type=sec_type, order_index=order,
                      time_limit_s=time_limit, config=config)
    db.session.add(section)
    db.session.commit()
    flash(f"{sec_type.title()} section added.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/delete", methods=["POST"])
@login_required
@teacher_required
def section_delete(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    db.session.delete(section)
    db.session.commit()
    flash("Section deleted.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/questions", methods=["POST"])
@login_required
@teacher_required
def question_add(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    q_type = request.form.get("type", QuestionType.MCQ)
    prompt = request.form.get("prompt", "").strip()
    correct = request.form.get("correct_answer", "").strip()
    group_id = request.form.get("group_id", "").strip() or None

    # Build MCQ options from option_a … option_d inputs
    options = None
    if q_type == QuestionType.MCQ:
        options = []
        for letter in ("A", "B", "C", "D"):
            text = request.form.get(f"option_{letter}", "").strip()
            if text:
                options.append({"id": letter, "text": text})

    order = len(section.questions) + 1
    question = Question(
        section_id=section_id,
        order_index=order,
        type=q_type,
        prompt=prompt,
        correct_answer=correct or None,
        options=options,
        group_id=group_id,
    )
    db.session.add(question)
    db.session.commit()
    flash("Question added.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/questions/<question_id>/delete",
                methods=["POST"])
@login_required
@teacher_required
def question_delete(exam_id: str, section_id: str, question_id: str):
    question = Question.query.filter_by(id=question_id, section_id=section_id).first_or_404()
    db.session.delete(question)
    db.session.commit()
    flash("Question deleted.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/questions/import", methods=["POST"])
@login_required
@teacher_required
def questions_import(exam_id: str, section_id: str):
    import csv
    import io
    from openpyxl import load_workbook

    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()

    upload = request.files.get("import_file")
    if not upload or not upload.filename:
        flash("No file selected.", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))

    fname = upload.filename.lower()
    rows = []
    try:
        if fname.endswith(".xlsx"):
            wb = load_workbook(upload, data_only=True)
            ws = wb.active
            raw_headers = next(ws.iter_rows(min_row=1, max_row=1, values_only=True))
            headers = [str(h or "").strip() for h in raw_headers]
            for row_vals in ws.iter_rows(min_row=2, values_only=True):
                rows.append({h: str(v or "").strip() for h, v in zip(headers, row_vals)})
        else:
            content = upload.read().decode("utf-8-sig")
            reader = csv.DictReader(io.StringIO(content))
            rows = [{k: (v or "").strip() for k, v in r.items()} for r in reader]
    except Exception as exc:
        flash(f"Could not parse file: {exc}", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))

    order = len(section.questions) + 1
    imported = 0
    errors = []

    for row_num, row in enumerate(rows, start=2):
        q_type = row.get("type", "").upper()
        prompt = row.get("prompt", "")
        if not q_type or not prompt:
            errors.append(f"Row {row_num}: 'type' and 'prompt' are required — skipped.")
            continue

        correct = row.get("correct_answer") or None
        group_id = row.get("group_id") or None
        try:
            marks = int(row.get("marks") or 1)
        except ValueError:
            marks = 1

        options = None
        if q_type in ("MCQ", "MATCHING_HEADINGS", "MATCHING_FEATURES"):
            opts = []
            for letter in "ABCDEFGH":
                text = row.get(f"option_{letter}", "")
                if text:
                    opts.append({"id": letter, "text": text})
            options = opts or None

        db.session.add(Question(
            section_id=section_id,
            order_index=order,
            type=q_type,
            prompt=prompt,
            correct_answer=correct,
            options=options,
            group_id=group_id,
            marks=marks,
        ))
        order += 1
        imported += 1

    db.session.commit()

    for err in errors[:5]:
        flash(err, "error")
    if imported:
        flash(f"{imported} question(s) imported.", "success")

    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


# ---------------------------------------------------------------------------
# Section content — Listening parts, Reading passages, Writing prompts
# ---------------------------------------------------------------------------

@admin_bp.route("/exams/<exam_id>/sections/<section_id>/upload-audio", methods=["POST"])
@login_required
@teacher_required
def upload_audio(exam_id: str, section_id: str):
    """Upload an audio file to R2 and return the resulting object key as JSON."""
    Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    audio = request.files.get("audio")
    if not audio or not audio.filename:
        return jsonify({"error": "No file provided."}), 400

    ext = audio.filename.rsplit(".", 1)[-1].lower() if "." in audio.filename else "mp3"
    allowed = {"mp3", "wav", "ogg", "m4a", "aac", "webm", "flac"}
    if ext not in allowed:
        return jsonify({"error": f"Unsupported format: {ext}"}), 400

    file_key = f"listening/{exam_id}/{section_id}/{uuid.uuid4()}.{ext}"
    try:
        from ..services.storage import upload_fileobj
        upload_fileobj(audio.stream, file_key, audio.content_type or f"audio/{ext}")
    except Exception as e:
        return jsonify({"error": str(e)}), 500

    return jsonify({"ok": True, "key": file_key})


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/parts", methods=["POST"])
@login_required
@teacher_required
def listening_part_add(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    audio_key = request.form.get("audio_key", "").strip()
    label = request.form.get("label", "").strip()
    group_id = request.form.get("group_id", "").strip()
    if not audio_key:
        flash("Audio file key is required.", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))
    config = dict(section.config)
    parts = list(config.get("parts", []))
    idx = len(parts) + 1
    parts.append({
        "audioFileKey": audio_key,
        "label": label or f"Part {idx}",
        "groupId": group_id or f"part-{idx}",
    })
    config["parts"] = parts
    section.config = config
    db.session.commit()
    flash("Audio part added.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/parts/<int:part_idx>/delete",
                methods=["POST"])
@login_required
@teacher_required
def listening_part_delete(exam_id: str, section_id: str, part_idx: int):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    config = dict(section.config)
    parts = list(config.get("parts", []))
    if 0 <= part_idx < len(parts):
        parts.pop(part_idx)
        config["parts"] = parts
        section.config = config
        db.session.commit()
        flash("Audio part removed.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/passages", methods=["POST"])
@login_required
@teacher_required
def reading_passage_add(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    title = request.form.get("title", "").strip()
    text = request.form.get("text", "").strip()
    group_id = request.form.get("group_id", "").strip()
    if not text:
        flash("Passage text is required.", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))
    config = dict(section.config)
    passages = list(config.get("passages", []))
    idx = len(passages) + 1
    passages.append({
        "title": title or f"Passage {idx}",
        "text": text,
        "groupId": group_id or f"passage-{idx}",
    })
    config["passages"] = passages
    section.config = config
    db.session.commit()
    flash("Passage added.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/passages/<int:passage_idx>/delete",
                methods=["POST"])
@login_required
@teacher_required
def reading_passage_delete(exam_id: str, section_id: str, passage_idx: int):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    config = dict(section.config)
    passages = list(config.get("passages", []))
    if 0 <= passage_idx < len(passages):
        passages.pop(passage_idx)
        config["passages"] = passages
        section.config = config
        db.session.commit()
        flash("Passage removed.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/writing-prompts", methods=["POST"])
@login_required
@teacher_required
def writing_prompts_update(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    config = dict(section.config)
    config["task1Prompt"] = request.form.get("task1Prompt", "").strip()
    config["task2Prompt"] = request.form.get("task2Prompt", "").strip()
    task1_image = request.form.get("task1ImageKey", "").strip()
    if task1_image:
        config["task1ImageKey"] = task1_image
    elif "task1ImageKey" in config and not task1_image:
        config.pop("task1ImageKey", None)
    section.config = config
    db.session.commit()
    flash("Writing prompts updated.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/speaking-parts", methods=["POST"])
@login_required
@teacher_required
def speaking_part_add(exam_id: str, section_id: str):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    question_text = request.form.get("question", "").strip()
    if not question_text:
        flash("Question text is required.", "error")
        return redirect(url_for("admin.exam_edit", exam_id=exam_id))
    config = dict(section.config)
    parts = list(config.get("parts", []))
    parts.append({"question": question_text})
    config["parts"] = parts
    section.config = config
    db.session.commit()
    flash("Speaking part added.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/exams/<exam_id>/sections/<section_id>/speaking-parts/<int:part_idx>/delete",
                methods=["POST"])
@login_required
@teacher_required
def speaking_part_delete(exam_id: str, section_id: str, part_idx: int):
    section = Section.query.filter_by(id=section_id, exam_id=exam_id).first_or_404()
    config = dict(section.config)
    parts = list(config.get("parts", []))
    if 0 <= part_idx < len(parts):
        parts.pop(part_idx)
        config["parts"] = parts
        section.config = config
        db.session.commit()
        flash("Speaking part removed.", "success")
    return redirect(url_for("admin.exam_edit", exam_id=exam_id))


@admin_bp.route("/sessions/<session_id>/reset", methods=["POST"])
@login_required
@teacher_required
def reset_session(session_id: str):
    """Reset a stuck SUBMITTED session back to IN_PROGRESS so the student can retry."""
    session = ExamSession.query.filter_by(id=session_id).first_or_404()
    if session.status != SessionStatus.SUBMITTED:
        flash(f"Session is in '{session.status}' state — only SUBMITTED sessions can be reset.", "error")
        return redirect(request.referrer or url_for("admin.dashboard"))
    session.status = SessionStatus.IN_PROGRESS
    db.session.commit()
    flash(f"Session reset to IN_PROGRESS. The student can now retake the speaking section.", "success")
    return redirect(request.referrer or url_for("admin.dashboard"))


@admin_bp.route("/sessions/<session_id>/retry-scoring", methods=["POST"])
@login_required
@teacher_required
def retry_scoring(session_id: str):
    """Re-enqueue AI scoring for a SUBMITTED session (e.g. after fixing Gemini billing)."""
    session = ExamSession.query.filter_by(id=session_id).first_or_404()
    if session.status != SessionStatus.SUBMITTED:
        flash(f"Session must be in SUBMITTED state to retry scoring (currently '{session.status}').", "error")
        return redirect(request.referrer or url_for("admin.dashboard"))
    from ..services.exam_engine import enqueue_scoring
    try:
        enqueue_scoring(session_id)
        flash("Scoring re-enqueued. Refresh the results page in a few seconds.", "success")
    except Exception as e:
        flash(f"Scoring failed: {e}", "error")
    return redirect(request.referrer or url_for("admin.dashboard"))


@admin_bp.route("/grading")
@login_required
@teacher_required
def grading_queue():
    pending = (
        ExamSession.query
        .filter(ExamSession.status.in_([SessionStatus.SUBMITTED, SessionStatus.SCORED]))
        .order_by(ExamSession.submitted_at.asc())
        .all()
    )
    return render_template("admin/grading.html", sessions=pending)


@admin_bp.route("/grading/<session_id>", methods=["GET", "POST"])
@login_required
@teacher_required
def grade_session(session_id: str):
    session = ExamSession.query.filter_by(id=session_id).first_or_404()
    scores = {s.section_type: s for s in session.scores}

    if request.method == "POST":
        for section_type, score in scores.items():
            override = request.form.get(f"override_{section_type}")
            feedback = request.form.get(f"feedback_{section_type}")
            if override:
                try:
                    score.teacher_override_score = float(override)
                except ValueError:
                    flash(f"Invalid score for {section_type}: must be a number.", "error")
                    return redirect(url_for("admin.grade_session", session_id=session_id))
            if feedback:
                score.teacher_feedback = feedback
            score.is_finalized = True
        db.session.commit()
        flash("Scores saved.", "success")
        return redirect(url_for("admin.grading_queue"))

    writing_responses = {r.task_number: r for r in session.writing_responses}
    speaking_responses = sorted(session.speaking_responses, key=lambda r: r.part_number)

    speaking_audio_urls = {}
    for resp in speaking_responses:
        if resp.audio_file_key:
            try:
                from ..services.storage import get_signed_download_url
                speaking_audio_urls[resp.part_number] = get_signed_download_url(
                    resp.audio_file_key, expires_in=3600
                )
            except Exception:
                pass

    return render_template(
        "admin/grade_session.html",
        session=session,
        scores=scores,
        writing_responses=writing_responses,
        speaking_responses=speaking_responses,
        speaking_audio_urls=speaking_audio_urls,
    )
